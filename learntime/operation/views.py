from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.utils.six import BytesIO
from django.views.generic import CreateView, DeleteView, DetailView
from django.views.generic.base import View

from learntime.operation.models import Log, FeedBack
from learntime.utils.helpers import PaginatorListView, RoleRequiredMixin, RootRequiredMixin
from learntime.users.enums import RoleEnum
from learntime.operation.tasks import send_email_to_user
from learntime.student.models import StudentCreditVerify


class LogList(LoginRequiredMixin, PaginatorListView):
    """日志列表页"""
    template_name = "operation/log_list.html"
    context_object_name = "logs"
    paginate_by = 50

    def get_queryset(self):
        if self.request.user.role == 1 or self.request.user.role == 2:
            return Log.objects.all().select_related("user")
        else:
            return Log.objects.filter(user=self.request.user).select_related("user")


class LogDetail(LoginRequiredMixin, DetailView):
    """日志详情页"""
    template_name = "operation/log_detail.html"
    context_object_name = "log"
    model = Log



class StudentActivityListView(LoginRequiredMixin, PaginatorListView):
    """学生参加活动列表页"""
    template_name = "operation/student_activity_list.html"
    context_object_name = "objects"
    paginate_by = 50

    def get_queryset(self):
        if self.request.user.role == RoleEnum.ROOT.value or self.request.user.role == RoleEnum.SCHOOL.value: # ROOT级别能看到所有记录
            return StudentCreditVerify.objects.all()
        else:
            return StudentCreditVerify.objects.filter(
                academy=self.request.user.academy
            )


class StudentActivityExportView(RoleRequiredMixin, View):
    """学生参加记录导出"""
    role_required = (RoleEnum.SCHOOL.value, RoleEnum.ROOT.value, RoleEnum.ACADEMY.value, RoleEnum.ORG.value)
    def get(self, request):
        import openpyxl
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet(title="data", index=0)

        # 写标题栏
        sheet.cell(1, 1, '活动名称')
        sheet.cell(1, 2, '主办方')
        sheet.cell(1, 3, '姓名')
        sheet.cell(1, 4, '学号')
        sheet.cell(1, 5, '学院')
        sheet.cell(1, 6, '班级')
        sheet.cell(1, 7, '参加类型')
        sheet.cell(1, 8, '获奖情况')
        sheet.cell(1, 9, '认定项目')
        sheet.cell(1, 10, '认定活动时')
        sheet.cell(1, 11, '填报人及联系方式')
        sheet.cell(1, 12, '审核人')
        sheet.cell(1, 13, '备注')
        sheet.cell(1, 14, '归属年度(如“2020-2021学年”)')


        student_activities = StudentCreditVerify.objects.all()

        if self.request.user.role == RoleEnum.ACADEMY.value:
            student_activities = student_activities.filter(
                academy=self.request.user.academy,
                grade=self.request.user.grade)
        if self.request.user.role == RoleEnum.ORG.value:
            student_activities = student_activities.filter(
                academy=self.request.user.academy
            )

        # 写数据
        row = 2
        for obj in student_activities: # 单条写入学生数据
            sheet.cell(row, 1, obj.activity_name)
            sheet.cell(row, 2, obj.sponsor)
            sheet.cell(row, 3, obj.name)
            sheet.cell(row, 4, obj.uid)
            sheet.cell(row, 5, obj.academy)
            sheet.cell(row, 6, obj.clazz)
            sheet.cell(row, 7, obj.join_type)
            sheet.cell(row, 8, obj.award)
            sheet.cell(row, 9, obj.credit_type)
            sheet.cell(row, 10, obj.credit)
            sheet.cell(row, 11, obj.contact)
            sheet.cell(row, 12, obj.to_name)
            sheet.cell(row, 13, obj.remark)
            sheet.cell(row, 14, obj.year)
            row += 1

        sio = BytesIO() # StringIO报错，使用BytesIO
        workbook.save(sio)
        sio.seek(0) # 定位到开始
        response = HttpResponse(sio.getvalue(), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=activity.xlsx'
        response.write(sio.getvalue())

        # 写入日志
        Log.objects.create(
            user=self.request.user,
            content=f"下载了{row - 1}条学生活动记录"
        )

        return response



class SearchRecordByPkAndTypeAPIView(LoginRequiredMixin, View):
    """查找某个学生某个类型的参加活动记录"""
    def get(self, request):
        """
        :param request:
        :param student_id: 学号
        :param type_name: 参与的类型
        :return:
        """
        student_id = request.GET.get("student_id")
        type_name = request.GET.get("type_name")
        if not student_id:
            return JsonResponse({"status": "fail", "reason": "学号查找不到"})

        if not type_name:
            records = StudentCreditVerify.objects.filter(uid=student_id)
        else:
            records = StudentCreditVerify.objects.filter(uid=student_id,
                                       credit_type=type_name)
        results = []
        for record in records:
            results.append({
                "activity_name": record.activity_name,
                "join_type": record.join_type,
                "credit": record.credit,
                "credit_type": record.credit_type,
                "student_name": record.name,
                "year": record.year,
                "to_name": record.to_name,
                "create_time": record.created_at.strftime("%Y-%m-%d"),
                "status": "签退成功"
            })

        return JsonResponse({"status": "ok", "data": results})


class FeedBackListView(LoginRequiredMixin, PaginatorListView):
    """反馈列表"""
    paginate_by = 20
    context_object_name = "feedbacks"
    template_name = "operation/feedback_list.html"

    def get_queryset(self):
        if self.request.user.role == RoleEnum.ROOT.value:
            # 管理员
            return FeedBack.objects.all()
        else:
            return FeedBack.objects.filter(email=self.request.user.email)


class FeedBackCreateView(RoleRequiredMixin, CreateView):
    role_required = (RoleEnum.SCHOOL.value, RoleEnum.ACADEMY.value, RoleEnum.STUDENT.value)
    template_name = "operation/feedback_create.html"
    fields = ('content', )
    model = FeedBack

    def form_valid(self, form):
        form.instance.name = self.request.user.name
        form.instance.email = self.request.user.email
        return super().form_valid(form)

    def get_success_url(self):
        messages.success(self.request, "提交反馈成功！")
        return reverse("operations:feedback_list")


class FeedBackDeleteView(LoginRequiredMixin, DeleteView):
    template_name = "operation/feedback_delete.html"
    model = FeedBack
    context_object_name = "feedback"
    def get_success_url(self):
        messages.success(self.request, "删除反馈成功！")
        return reverse("operations:feedback_list")


class FeedBackDetailAPIView(LoginRequiredMixin, View):
    """反馈详情api接口"""
    def get(self, request):
        try:
            feedback_id = request.GET.get('feedback_id')
            feedback = FeedBack.objects.get(pk=feedback_id)
            return JsonResponse({
                "status": "ok",
                "content": feedback.content,
                "name": feedback.name
            })
        except Exception:
            return JsonResponse({"status": "fail"})


class SendEmailAPIView(RootRequiredMixin, View):
    """发送邮件接口"""
    def post(self, request):
        try:
            title = request.POST.get("title")
            content = request.POST.get("content")
            email = request.POST.get("email")
            feedback_id = request.POST.get("feedback_id")

            # 修改数据库记录
            feedback = FeedBack.objects.get(pk=feedback_id)
            feedback.is_fix = True
            feedback.reply = content
            feedback.save()
            send_email_to_user(email, title, content)

        except Exception:
            return JsonResponse({"status": "fail"})
        return JsonResponse({"status": "ok"})
